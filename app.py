import openpyxl
from openpyxl import Workbook
from flask import Flask, render_template, request, flash, redirect, url_for

app = Flask(__name__)
app.secret_key = "your_secret_key"

store = "responses.xlsx"

def email_exists(email):
    try:
        workbook = openpyxl.load_workbook(store)
        sheet = workbook.active
        m_row = sheet.max_row

        for i in range(2, m_row + 1): 
            cell_value = sheet.cell(row=i, column=1).value
            if cell_value == email:
                return True
    except FileNotFoundError:
        return False
    return False

def save_response(data):
    try:
        workbook = openpyxl.load_workbook(store)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["email", "question1", "question2", "question3", "question4", "question5"])

    sheet.append([
        data["email"], 
        data["question1"], 
        data["question2"], 
        data["question3"], 
        data["question4"], 
        data["question5"]
    ])
    
    workbook.save(store)

@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        email = request.form.get("email")
        if email_exists(email):
            flash("This email has already been used to submit a response.")
            return redirect(url_for("form"))
        
        responses = {
            "email": email,
            "question1": request.form.get("question1"),
            "question2": request.form.get("question2"),
            "question3": request.form.get("question3"),
            "question4": request.form.get("question4"),
            "question5": request.form.get("question5"),
        }
        save_response(responses)
        return redirect(url_for("thank_you"))

    return render_template("forms.html")

@app.route("/thank-you")
def thank_you():
    return render_template("thankyou.html")

if __name__ == "__main__":
    app.run(debug=True)
